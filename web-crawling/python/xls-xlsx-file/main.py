import openpyxl

wb = openpyxl.load_workbook('./member_test.xlsx')
sheet = wb["Sheet1"] # 불러온 Excel 파일의 Sheet1이라는 시트를 불러온다.

member = {}
for i in range(2, sheet.max_row + 1):
    name = sheet.cell(row=i, column=3).value
    number = sheet.cell(row=i, column=2).value
    # cell() 메서드의 인자로 row와 column에 값을 지정해주는데, row는 각 row(데이터)를 의미하고 column은 표의 컬럼을 의미하고 컬럼의 인덱스 번호를 입력한다.
    member[name] = number

print(member)